import re
import pandas as pd
from openpyxl import load_workbook

# Configuracion de rutas de archivos y nombres de columnas
ARCHIVO_ENTRADA = "Informacion_Usuarios.xls"
ARCHIVO_SALIDA  = "Informacion_Usuarios_telefonos_estandarizados.xlsx"
COLUMNA_TELEFONO = "Celular"

# Columnas que no son relevantes para el proceso de marcado pero se mantienen en el registro
COLUMNAS_OCULTAR = [
    "Apellido Materno",
    "Manager",
    "Concepto",
    "Idioma"
]

def normalize_candidates(cell):
    """
    Limpia y extrae numeros telefonicos de 10 digitos de una cadena de texto.
    Maneja multiples delimitadores y limpia prefijos de marcacion comunes en Mexico.
    """
    if cell is None:
        return []
    s = str(cell).strip()
    if not s:
        return []

    # Divide la cadena si contiene multiples telefonos separados por coma o punto y coma
    parts = [p.strip() for p in re.split(r"[;,]", s) if p.strip()]
    nums = []

    for p in parts:
        # Ignora valores nulos o no definidos explicitamente en el texto
        if p.lower() in ("n/a", "na", "null", "none", "s/n"):
            continue

        # Elimina cualquier caracter no numerico
        digits = re.sub(r"\D", "", p)

        # Reglas de normalizacion para estandarizar a 10 digitos (formato local MX)
        if digits.startswith("01152"):
            digits = digits[5:]
        if digits.startswith("01") and len(digits) > 10:
            digits = digits[2:]
        if digits.startswith("52") and len(digits) >= 12:
            digits = digits[-10:]
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) > 10:
            digits = digits[-10:]

        # Solo se aceptan numeros que cumplen con la longitud de la red nacional
        if len(digits) == 10:
            nums.append(digits)

    # Elimina duplicados manteniendo el orden original
    return list(dict.fromkeys(nums))

def main():
    """
    Funcion principal: Orquesta la lectura, procesamiento y formateo del archivo Excel.
    """
    # Lectura inicial forzando tipos string para evitar truncamiento de ceros a la izquierda
    df = pd.read_excel(ARCHIVO_ENTRADA, dtype=str)

    # Aplicacion de logica de limpieza en la columna de origen
    cands = df[COLUMNA_TELEFONO].apply(normalize_candidates)
    
    # Expansion de resultados a nuevas columnas de control
    df["Celular_10"] = cands.apply(lambda x: x[0] if x else None)
    df["Celular_Alt_10"] = cands.apply(lambda x: x[1] if len(x) > 1 else None)
    
    # Formato E164 simplificado para integracion directa con troncales SIP/Issabel
    df["Celular_E164_MX"] = df["Celular_10"].apply(
        lambda x: f"52{x}" if isinstance(x, str) and len(x) == 10 else None
    )

    # Persistencia inicial de datos procesados
    df.to_excel(ARCHIVO_SALIDA, index=False)

    # Post-procesamiento estético y funcional del archivo Excel mediante openpyxl
    wb = load_workbook(ARCHIVO_SALIDA)
    ws = wb.active

    # Ajuste de altura de filas para mejorar la legibilidad en pantalla
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 22

    # Calculo dinamico del ancho de columnas segun el contenido
    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max(max_length + 4, 15), 45)

    # Mapeo de encabezados para aplicar ocultamiento de columnas especificas
    header = {cell.value: cell.column_letter for cell in ws[1]}
    for col_name in COLUMNAS_OCULTAR:
        col_letter = header.get(col_name)
        if col_letter:
            ws.column_dimensions[col_letter].hidden = True

    # Inmovilizacion del panel superior para facilitar la navegacion en bases grandes
    ws.freeze_panes = "A2"

    # Guardado final de la version estandarizada
    wb.save(ARCHIVO_SALIDA)
    print("Excel generado exitosamente: Procesamiento de estandarizacion finalizado.")

if __name__ == "__main__":
    main()