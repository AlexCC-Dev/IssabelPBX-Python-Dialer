import re
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

ARCHIVO_ENTRADA = "Informacion_Usuarios_telefonos_estandarizados.xlsx" 
COLUMNA_TELEFONO_BUSQUEDA = "Celular_10"

COLUMNAS_OCULTAR = ["Apellido Materno", "Manager", "Concepto", "Idioma"]

def normalize_phone10(v):
    """Convierte cualquier formato a 10 dígitos (si se puede)."""
    if v is None:
        return None

    digits = re.sub(r"\D", "", str(v).strip())
    if not digits:
        return None

    if digits.startswith("01152"):
        digits = digits[5:]  # quita 01152
    if digits.startswith("01") and len(digits) > 10:
        digits = digits[2:]  # quita 01
    if digits.startswith("52") and len(digits) >= 12:
        digits = digits[-10:]  # deja solo los últimos 10
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]  # quita 1 (NANP)
    if len(digits) > 10:
        digits = digits[-10:]  # último recurso

    return digits if len(digits) == 10 else None

def ajustar_excel(ws):
    """Hace el Excel más legible: ancho de columnas + alto de filas."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 22

    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max(max_length + 4, 15), 45)

def main():
    tel_input = input("📞 Escribe el teléfono a buscar (10 dígitos o con formato): ").strip()
    tel10 = normalize_phone10(tel_input)

    if not tel10:
        print("❌ Teléfono inválido. Prueba con 10 dígitos.")
        return

    df = pd.read_excel(ARCHIVO_ENTRADA, dtype=str)

    if COLUMNA_TELEFONO_BUSQUEDA not in df.columns:
        print(f"❌ No existe la columna '{COLUMNA_TELEFONO_BUSQUEDA}' en el archivo.")
        print(f"Columnas disponibles: {list(df.columns)}")
        return

    matches = df[df[COLUMNA_TELEFONO_BUSQUEDA].astype(str) == tel10]

    if matches.empty:
        print(f"❌ No se encontró el teléfono: {tel10}")
        return

    print(f"✅ Encontré {len(matches)} coincidencia(s) para {tel10}")

    salida = f"resultado_{tel10}.xlsx"
    df.to_excel(salida, index=False)

    wb = load_workbook(salida)
    ws = wb.active

    ajustar_excel(ws)

    fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    for idx in matches.index.tolist():
        excel_row = idx + 2
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = fill

    header = {cell.value: cell.column_letter for cell in ws[1]}
    for col_name in COLUMNAS_OCULTAR:
        col_letter = header.get(col_name)
        if col_letter:
            ws.column_dimensions[col_letter].hidden = True
    ws.freeze_panes = "A2"

    wb.save(salida)

    print(f"📂 Abriendo: {salida}")
    os.startfile(salida)

if __name__ == "__main__":
    main()
